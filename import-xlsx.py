#!/usr/bin/env python3
"""
Import BeFine article data from the provided Excel workbook into Supabase.

Dependencies:
  pip install openpyxl supabase

Run:
  python3 import-xlsx.py
"""

from __future__ import annotations

from decimal import Decimal, InvalidOperation
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook
from supabase import create_client


INPUT_FILE = "/Users/suzukiokiichi/Downloads/BeFineClinic .xlsx"
SUPABASE_URL = "https://hwfbpissyywzqbdthpek.supabase.co"
SUPABASE_KEY = "sb_publishable_kzzmDG971X8Vvr2w_RPsPw_MW-o_104"
TABLE_NAME = "articles"
BATCH_SIZE = 100

DIRECTORY_OFFSETS = {
    "美容医療": 0,
    "医療脱毛": 10000,
    "クリニック": 20000,
    "脱毛器": 30000,
    "ダイエット解説記事": 40000,
}

COLUMN_MAPPING = {
    "ID": "id",
    "keyword": "keyword",
    "genre": "genre",
    "category": "category",
    "release": "release_status",
    "designer_release": "designer_release",
    "blog_parts": "blog_parts",
    "No.1_client": "no1_client",
    "seasonal_offer": "seasonal_appeal",
    "NG_keyword": "banned_keywords",
    "additional-information": "notes",
    "survey_preparation": "survey_status",
    "survey_url": "survey_url",
    "figma_url": "figma_url",
    "picture_url": "image_folder_url",
    "emergent_information": "urgent_notice",
    "regulation": "regulation",
    "title": "title",
    "url-link": "url",
    "url-draft": "draft_url",
    "movie": "movie",
    "meta_discription": "meta_description",
    "picture_request_availablity": "picture_request",
}


def normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and value.strip() == "")


def parse_excel_id(value: Any) -> Optional[int]:
    if is_blank(value):
        return None

    if isinstance(value, bool):
        raise ValueError(f"Invalid boolean ID value: {value!r}")

    if isinstance(value, int):
        return value

    if isinstance(value, float):
        if not value.is_integer():
            raise ValueError(f"Invalid non-integer numeric ID value: {value!r}")
        return int(value)

    try:
        decimal_value = Decimal(str(value).strip())
    except (InvalidOperation, ValueError) as exc:
        raise ValueError(f"Invalid ID value: {value!r}") from exc

    if decimal_value != decimal_value.to_integral_value():
        raise ValueError(f"Invalid non-integer ID value: {value!r}")

    return int(decimal_value)


def clean_cell(value: Any) -> Any:
    if isinstance(value, str):
        stripped = value.strip()
        return stripped if stripped else None
    return value


def send_batch(client: Any, batch: List[Dict[str, Any]], sheet_name: str, imported: int) -> None:
    if not batch:
        return

    print(f"[{sheet_name}] Upserting batch of {len(batch)} rows; total prepared: {imported}")
    response = client.table(TABLE_NAME).upsert(batch, on_conflict="id").execute()
    if getattr(response, "data", None) is None and getattr(response, "error", None):
        raise RuntimeError(f"Supabase upsert failed on sheet {sheet_name}: {response.error}")
    print(f"[{sheet_name}] Upserted batch successfully")


def build_header_index(sheet: Any) -> Dict[str, int]:
    field_row = next(sheet.iter_rows(min_row=2, max_row=2, values_only=True), None)
    if field_row is None:
        raise ValueError(f"Sheet {sheet.title} does not have row 2 field names")

    return {
        normalize_header(value): index
        for index, value in enumerate(field_row)
        if normalize_header(value)
    }


def import_sheet(client: Any, sheet: Any) -> int:
    sheet_name = sheet.title
    offset = DIRECTORY_OFFSETS[sheet_name]
    header_index = build_header_index(sheet)

    if "ID" not in header_index:
        raise ValueError(f"Sheet {sheet_name} is missing required row 2 field: ID")

    missing_fields = [field for field in COLUMN_MAPPING if field not in header_index]
    if missing_fields:
        print(f"[{sheet_name}] Warning: missing mapped fields: {', '.join(missing_fields)}")

    batch: List[Dict[str, Any]] = []
    imported = 0
    skipped_empty_id = 0

    print(f"[{sheet_name}] Starting import with directory offset {offset}")

    for excel_row_number, row in enumerate(sheet.iter_rows(min_row=3, values_only=True), start=3):
        raw_id = row[header_index["ID"]] if header_index["ID"] < len(row) else None
        local_id = parse_excel_id(raw_id)
        if local_id is None:
            skipped_empty_id += 1
            continue

        record: Dict[str, Any] = {"directory": sheet_name}
        for excel_field, supabase_column in COLUMN_MAPPING.items():
            if excel_field not in header_index:
                continue

            column_index = header_index[excel_field]
            value = row[column_index] if column_index < len(row) else None

            if excel_field == "ID":
                record[supabase_column] = local_id + offset
            else:
                record[supabase_column] = clean_cell(value)

        batch.append(record)
        imported += 1

        if len(batch) >= BATCH_SIZE:
            send_batch(client, batch, sheet_name, imported)
            batch.clear()

        if imported % 500 == 0:
            print(f"[{sheet_name}] Prepared {imported} rows through Excel row {excel_row_number}")

    send_batch(client, batch, sheet_name, imported)
    print(
        f"[{sheet_name}] Finished: imported {imported} rows, "
        f"skipped {skipped_empty_id} rows with empty ID"
    )
    return imported


def main() -> None:
    print(f"Loading workbook: {INPUT_FILE}")
    workbook = load_workbook(INPUT_FILE, read_only=True, data_only=True)
    client = create_client(SUPABASE_URL, SUPABASE_KEY)

    try:
        total_imported = 0
        for sheet_name in DIRECTORY_OFFSETS:
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"Workbook is missing expected sheet: {sheet_name}")

            total_imported += import_sheet(client, workbook[sheet_name])

        print(f"Import complete. Total imported rows: {total_imported}")
    finally:
        workbook.close()


if __name__ == "__main__":
    main()
